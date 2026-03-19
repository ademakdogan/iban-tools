"""Scraper for Luxembourg (LU) bank codes from ABBL.

Source: ABBL IBAN-BIC codes XLSX
URL:    https://abbl.lu/en/professionals/page/iban-and-bic-codes
"""

import json
import sys
from pathlib import Path

import openpyxl
import requests
from bs4 import BeautifulSoup

DATA_DIR = Path(__file__).parent.parent / "src" / "iban_tools" / "data"


def scrape_lu() -> dict[str, str]:
    """Scrape Luxembourg bank code → BIC mapping from ABBL."""
    list_url = "https://abbl.lu/en/professionals/page/iban-and-bic-codes"

    print("[LU] Fetching ABBL IBAN-BIC page...")
    page = requests.get(list_url, timeout=30)
    page.raise_for_status()

    soup = BeautifulSoup(page.text, "html.parser")

    xlsx_url = None
    for link in soup.find_all("a"):
        href = link.get("href", "")
        if "Luxembourg Register" in href and href.endswith(".xlsx"):
            xlsx_url = href
            break
        # Also check link text
        text = link.get_text()
        if "IBAN-BIC" in text and href.endswith(".xlsx"):
            xlsx_url = href
            break

    if not xlsx_url:
        # Fallback: search any xlsx link
        for link in soup.find_all("a"):
            href = link.get("href", "")
            if href.endswith(".xlsx"):
                xlsx_url = href
                break

    if not xlsx_url:
        print("[LU] ERROR: Could not find XLSX download link on ABBL page.")
        return {}

    if xlsx_url.startswith("/") and not xlsx_url.startswith("//"):
        xlsx_url = "https://abbl.lu" + xlsx_url

    print(f"[LU] Downloading XLSX from {xlsx_url}...")
    resp = requests.get(xlsx_url, timeout=60)
    resp.raise_for_status()

    wb = openpyxl.load_workbook(filename=__import__("io").BytesIO(resp.content), read_only=True)

    # Try "Organizations" sheet
    ws = wb["Organizations"] if "Organizations" in wb.sheetnames else wb.active

    bank_codes: dict[str, str] = {}
    # Data starts at row 3 (row 2 is header)
    rows = list(ws.iter_rows(min_row=3, values_only=True))

    for row in rows:
        if not row or len(row) < 3:
            continue

        iban_code = str(row[1]).strip() if row[1] else ""
        bic_code = str(row[2]).strip().replace(" ", "") if row[2] else ""

        if iban_code and bic_code:
            bank_codes[iban_code] = bic_code

    wb.close()
    print(f"[LU] Found {len(bank_codes)} bank codes.")
    return bank_codes


def main():
    data = scrape_lu()
    if data:
        output = DATA_DIR / "lu.json"
        with output.open("w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False)
        print(f"[LU] Written to {output}")
    else:
        print("[LU] No data scraped, keeping existing file.", file=sys.stderr)


if __name__ == "__main__":
    main()
